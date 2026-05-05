import csv
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from csv_to_excel import _build_workbook, _convert_broker, _sheet_name


def _write_csv(path: Path, rows: list[list[str]]) -> None:
    with path.open('w', newline='', encoding='utf-8') as f:
        csv.writer(f).writerows(rows)


def _sheet_rows(sheet) -> list[list]:
    return [list(row) for row in sheet.iter_rows(values_only=True)]


class TestCsvToExcel:
    # ------------------------------------------------------------------
    # _sheet_name
    # ------------------------------------------------------------------

    def test_year_prefix_stripped_from_sheet_name(self) -> None:
        assert _sheet_name(Path('2024_transactions.csv')) == 'transactions'

    def test_year_prefix_stripped_leaving_compound_name(self) -> None:
        assert _sheet_name(Path('2024_gains_2074.csv')) == 'gains_2074'

    def test_sheet_name_kept_when_no_year_prefix(self) -> None:
        assert _sheet_name(Path('summary.csv')) == 'summary'

    def test_sheet_name_truncated_to_31_chars(self) -> None:
        long_name = Path(f'2024_{"x" * 40}.csv')
        assert len(_sheet_name(long_name)) <= 31

    # ------------------------------------------------------------------
    # _build_workbook — sheet structure
    # ------------------------------------------------------------------

    def test_single_csv_produces_one_sheet(self, tmp_path: Path) -> None:
        _write_csv(tmp_path / '2024_transactions.csv', [['Date', 'ISIN']])

        workbook = _build_workbook(tmp_path)

        assert workbook.sheetnames == ['transactions']

    def test_multiple_csvs_produce_multiple_sheets(self, tmp_path: Path) -> None:
        _write_csv(tmp_path / '2024_dividendes.csv', [['Date', 'Montant']])
        _write_csv(tmp_path / '2024_transactions.csv', [['Date', 'ISIN']])

        workbook = _build_workbook(tmp_path)

        assert set(workbook.sheetnames) == {'dividendes', 'transactions'}

    def test_sheets_ordered_alphabetically_by_csv_filename(self, tmp_path: Path) -> None:
        _write_csv(tmp_path / '2024_transactions.csv', [['H']])
        _write_csv(tmp_path / '2024_dividendes.csv', [['H']])
        _write_csv(tmp_path / '2024_summary.csv', [['H']])

        workbook = _build_workbook(tmp_path)

        assert workbook.sheetnames == ['dividendes', 'summary', 'transactions']

    def test_empty_directory_produces_workbook_with_no_sheets(self, tmp_path: Path) -> None:
        workbook = _build_workbook(tmp_path)

        assert workbook.sheetnames == []

    # ------------------------------------------------------------------
    # _build_workbook — content fidelity
    # ------------------------------------------------------------------

    def test_sheet_content_matches_csv_rows_exactly(self, tmp_path: Path) -> None:
        rows = [['Date', 'ISIN', 'Montant EUR'], ['2024-01-15', 'IE00B41N0724', '100.00']]
        _write_csv(tmp_path / '2024_transactions.csv', rows)

        sheet = _build_workbook(tmp_path)['transactions']

        assert _sheet_rows(sheet) == rows

    def test_each_sheet_matches_its_own_csv(self, tmp_path: Path) -> None:
        div_rows = [['Date', 'Montant'], ['2024-06-01', '50.00']]
        tx_rows = [['Date', 'ISIN'], ['2024-01-15', 'LU0852473015']]
        _write_csv(tmp_path / '2024_dividendes.csv', div_rows)
        _write_csv(tmp_path / '2024_transactions.csv', tx_rows)

        workbook = _build_workbook(tmp_path)

        assert _sheet_rows(workbook['dividendes']) == div_rows
        assert _sheet_rows(workbook['transactions']) == tx_rows

    def test_row_count_matches_csv(self, tmp_path: Path) -> None:
        rows = [['H1', 'H2']] + [[f'r{i}a', f'r{i}b'] for i in range(10)]
        _write_csv(tmp_path / '2024_summary.csv', rows)

        sheet = _build_workbook(tmp_path)['summary']

        assert len(_sheet_rows(sheet)) == len(rows)

    def test_column_count_matches_csv(self, tmp_path: Path) -> None:
        rows = [['A', 'B', 'C', 'D', 'E'], ['1', '2', '3', '4', '5']]
        _write_csv(tmp_path / '2024_fx_log.csv', rows)

        sheet = _build_workbook(tmp_path)['fx_log']

        for sheet_row, csv_row in zip(_sheet_rows(sheet), rows):
            assert len(sheet_row) == len(csv_row)

    def test_accented_characters_preserved(self, tmp_path: Path) -> None:
        rows = [['Titre', 'Taux BCE', 'Montant arrondi'], ['fonds d\'actions', '1.0000', '1 234 €']]
        _write_csv(tmp_path / '2024_dividendes.csv', rows)

        sheet = _build_workbook(tmp_path)['dividendes']

        assert _sheet_rows(sheet) == rows

    def test_header_row_is_first_sheet_row(self, tmp_path: Path) -> None:
        header = ['Date cession', 'ISIN', 'Plus/moins-value EUR (PMP)']
        rows = [header, ['2024-03-10', 'GB00BJYDH287', '+500.00']]
        _write_csv(tmp_path / '2024_gains_2074.csv', rows)

        sheet = _build_workbook(tmp_path)['gains_2074']

        assert _sheet_rows(sheet)[0] == header

    # ------------------------------------------------------------------
    # _convert_broker
    # ------------------------------------------------------------------

    def test_convert_broker_creates_xlsx_at_expected_path(self, tmp_path: Path) -> None:
        year_dir = tmp_path / '2024'
        broker_dir = year_dir / 'yuh'
        broker_dir.mkdir(parents=True)
        excel_dir = year_dir / 'excel'
        excel_dir.mkdir()
        _write_csv(broker_dir / '2024_transactions.csv', [['Date'], ['2024-01-01']])

        _convert_broker('yuh', 2024, year_dir, excel_dir)

        assert (excel_dir / 'yuh_2024_ifu.xlsx').exists()

    def test_convert_broker_returns_true_on_success(self, tmp_path: Path) -> None:
        year_dir = tmp_path / '2024'
        broker_dir = year_dir / 'yuh'
        broker_dir.mkdir(parents=True)
        excel_dir = year_dir / 'excel'
        excel_dir.mkdir()
        _write_csv(broker_dir / '2024_transactions.csv', [['Date'], ['2024-01-01']])

        result = _convert_broker('yuh', 2024, year_dir, excel_dir)

        assert result is True

    def test_convert_broker_returns_false_when_broker_dir_missing(self, tmp_path: Path) -> None:
        year_dir = tmp_path / '2024'
        year_dir.mkdir()
        excel_dir = year_dir / 'excel'
        excel_dir.mkdir()

        result = _convert_broker('yuh', 2024, year_dir, excel_dir)

        assert result is False

    def test_convert_broker_returns_false_when_no_csvs_in_dir(self, tmp_path: Path) -> None:
        year_dir = tmp_path / '2024'
        broker_dir = year_dir / 'yuh'
        broker_dir.mkdir(parents=True)
        excel_dir = year_dir / 'excel'
        excel_dir.mkdir()

        result = _convert_broker('yuh', 2024, year_dir, excel_dir)

        assert result is False

    def test_convert_broker_excel_content_matches_source_csv(self, tmp_path: Path) -> None:
        year_dir = tmp_path / '2024'
        broker_dir = year_dir / 'yuh'
        broker_dir.mkdir(parents=True)
        excel_dir = year_dir / 'excel'
        excel_dir.mkdir()
        rows = [['Ticker', 'ISIN', 'Montant EUR'], ['MSFT', 'US5949181045', '999.99']]
        _write_csv(broker_dir / '2024_transactions.csv', rows)

        _convert_broker('yuh', 2024, year_dir, excel_dir)

        workbook = openpyxl.load_workbook(excel_dir / 'yuh_2024_ifu.xlsx')
        assert _sheet_rows(workbook['transactions']) == rows
